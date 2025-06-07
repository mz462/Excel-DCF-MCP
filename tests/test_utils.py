import unittest
from excel_mcp.utils import (
    address_within_ranges,
    collect_column_outputs,
    gather_row_outputs,
    refine_header_cells,
)
from openpyxl import Workbook


class TestAddressWithinRanges(unittest.TestCase):
    def test_address_inside(self):
        ranges = ["Sheet1!A1:C3", "Sheet2!D4:E5"]
        self.assertTrue(address_within_ranges("Sheet1!B2", ranges))

    def test_address_outside(self):
        ranges = ["Sheet1!A1:C3"]
        self.assertFalse(address_within_ranges("Sheet1!D1", ranges))

    def test_sheet_mismatch(self):
        ranges = ["Sheet1!A1:C3"]
        self.assertFalse(address_within_ranges("Sheet2!A1", ranges))


class TestCollectColumnOutputs(unittest.TestCase):
    def test_basic_scan(self):
        cells = {
            "A1": {"output": "Header"},
            "A2": {"output": 5},
            "A3": {"output": "x", "formula": "=B1"},
            "A4": {"output": 10},
            "A5": {"output": "stop"},
        }

        result = collect_column_outputs(cells, "A4", text_limit=1)
        expected = {
            "A3": "x",
            "A2": 5,
            "A1": "Header",
            "A4": 10,
            "A5": "stop",
        }
        self.assertEqual(result, expected)


class TestGatherRowOutputs(unittest.TestCase):
    def test_basic_scan(self):
        cells = {
            "A1": {"output": "Left"},
            "B1": {"output": 5},
            "C1": {"output": "x", "formula": "=A1"},
            "D1": {"output": 10},
            "E1": {"output": "stop"},
        }

        result = gather_row_outputs(cells, "D1", text_limit=1)
        expected = {
            "C1": "x",
            "B1": 5,
            "A1": "Left",
            "D1": 10,
            "E1": "stop",
        }
        self.assertEqual(result, expected)


class TestRefineHeaderCells(unittest.TestCase):
    def test_prune_headers_basic(self):
        wb = Workbook()
        ws = wb.active
        ws["B1"] = "Col"
        ws["A2"] = "Row"
        ws["B2"] = 42

        col_vals, row_vals, val = refine_header_cells(["B1", "C1"], ["A2", "A3"], "B2", ws)
        self.assertEqual(col_vals, ["Col"])
        self.assertEqual(row_vals, ["Row"])
        self.assertEqual(val, 42)


if __name__ == "__main__":
    unittest.main()
